# The program will receive a list of terraria weapons from the user,
# It will then automatically do a search for each of them in the terraria wiki,
# scrape their stats, and implement them in an Excel table,
# in which their stats will be calculated to conclude which weapon deals the most damage.

import openpyxl as xl
import webDrivers
import re
import time

# setting the excel file
wb = xl.load_workbook('Terraria_weapons.xlsx')
sheetRanged = wb['ranged']

# setting the web driver
url = "https://terraria.gamepedia.com/Ranged_weapons"
driver = webDrivers.WebDriver(url)

# Fills a list of weapons until user types "done"
weapon_list = []
print("Enter weapon name, 'done' to stop")
weapon = input()
while weapon != "done":
    weapon_list.append(weapon)
    weapon = input()

# Goes through the weapon list, scrape information for each weapon.
for wp in weapon_list:
    max = sheetRanged.max_row + 1
    driver.choose_weapon_by_name(wp)
    values = driver.get_info()

# put all the stats in variables, uses the re library to remove unwanted text/numbers.
    sheetRanged.cell(max,1).value = wp
    ammo = values[1].text
    damage = re.findall(r"[-+]?\d*\.\d+|\d+", values[2].text)[0]
    knockback = re.findall(r"[-+]?\d*\.\d+|\d+", values[3].text)[0]
    crit = re.findall(r"[-+]?\d*\.\d+|\d+", values[4].text)[0]
    use_time = re.findall(r"[-+]?\d*\.\d+|\d+", values[5].text)[0]
    velocity = re.findall(r"[-+]?\d*\.\d+|\d+", values[6].text)[0]
    effect = values[7].text

# implements the stats in the excel file
    for column in range(1, sheetRanged.max_column):

        cell = sheetRanged.cell(max,column)
        title = sheetRanged.cell(1,column).value
        if title == "Ammo":
            print("found ammo")
            cell.value = ammo
        elif title == "Damage":
            print("found damage")
            cell.value = damage
        elif title == "Knockback":
            print("found knockback")
            cell.value=knockback
        elif title == "Crit chance":
            print("found crit")
            cell.value = crit
        elif title == "Use time":
            print("found usetime")
            cell.value = use_time
        elif title == "Velocity":
            print("found velocity")
            cell.value = velocity
        elif title == "Effect":
            print("found effect")
            cell.value = effect
        elif title == "Projectiles":
            cell.value = 1
        else:
            pass

    driver.back()
    time.sleep(1)

wb.save('Terraria_weapons2.xlsx')
driver.quit()